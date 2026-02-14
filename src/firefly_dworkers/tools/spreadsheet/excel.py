"""Excel adapter for SpreadsheetPort."""

from __future__ import annotations

import asyncio
import io
import logging
from collections.abc import Sequence
from typing import Any

from fireflyframework_genai.tools.base import GuardProtocol

from firefly_dworkers.tools.registry import tool_registry
from firefly_dworkers.tools.spreadsheet.base import SpreadsheetPort
from firefly_dworkers.tools.spreadsheet.models import (
    SheetData,
    SheetSpec,
    SpreadsheetOperation,
    WorkbookData,
)

logger = logging.getLogger(__name__)

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _require_openpyxl() -> None:
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl required: pip install firefly-dworkers[data]")


@tool_registry.register("excel", category="spreadsheet")
class ExcelTool(SpreadsheetPort):
    """Read, create, and modify Excel (.xlsx) workbooks using openpyxl."""

    def __init__(
        self,
        *,
        timeout: float = 60.0,
        guards: Sequence[GuardProtocol] = (),
    ) -> None:
        super().__init__(
            "excel",
            description="Read, create, and modify Excel (.xlsx) workbooks.",
            timeout=timeout,
            guards=guards,
        )

    async def _read_spreadsheet(self, source: str, sheet_name: str = "") -> WorkbookData:
        _require_openpyxl()
        return await asyncio.to_thread(self._read_sync, source, sheet_name)

    async def _create_spreadsheet(self, sheets: list[SheetSpec]) -> bytes:
        _require_openpyxl()
        return await asyncio.to_thread(self._create_sync, sheets)

    async def _modify_spreadsheet(self, source: str, operations: list[SpreadsheetOperation]) -> bytes:
        _require_openpyxl()
        return await asyncio.to_thread(self._modify_sync, source, operations)

    # -- synchronous helpers ---------------------------------------------------

    def _read_sync(self, source: str, sheet_name: str) -> WorkbookData:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
        sheets: list[SheetData] = []

        target_sheets = [wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames else [wb.active]

        for ws in target_sheets:
            headers: list[str] = []
            rows: list[list[Any]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                else:
                    rows.append(list(row))

            sheets.append(
                SheetData(
                    name=ws.title or "",
                    headers=headers,
                    rows=rows,
                    row_count=len(rows),
                    col_count=len(headers),
                )
            )

        active_name = wb.active.title if wb.active else ""
        wb.close()

        return WorkbookData(
            sheets=sheets,
            active_sheet=active_name,
        )

    def _create_sync(self, sheets: list[SheetSpec]) -> bytes:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        # Remove the default sheet
        if wb.active:
            wb.remove(wb.active)

        for spec in sheets:
            ws = wb.create_sheet(title=spec.name)

            # -- Write headers --
            if spec.headers:
                ws.append(spec.headers)
                if spec.header_style:
                    font, fill, alignment = self._make_openpyxl_style(
                        spec.header_style, Font, PatternFill, Alignment,
                    )
                    for col_idx in range(1, len(spec.headers) + 1):
                        cell = ws.cell(row=1, column=col_idx)
                        if font:
                            cell.font = font
                        if fill:
                            cell.fill = fill
                        if alignment:
                            cell.alignment = alignment

            # -- Write data rows --
            data_start_row = 2 if spec.headers else 1
            for row in spec.rows:
                ws.append(row)

            # -- Apply cell_style to all data cells --
            if spec.cell_style and spec.rows:
                font, fill, alignment = self._make_openpyxl_style(
                    spec.cell_style, Font, PatternFill, Alignment,
                )
                max_col = max((len(r) for r in spec.rows), default=0)
                for row_idx in range(data_start_row, data_start_row + len(spec.rows)):
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if font:
                            cell.font = font
                        if fill:
                            cell.fill = fill
                        if alignment:
                            cell.alignment = alignment

            # -- Column widths --
            for i, width in enumerate(spec.column_widths):
                col_letter = get_column_letter(i + 1)
                ws.column_dimensions[col_letter].width = width

            # -- Number formats --
            if spec.number_formats:
                max_row = ws.max_row or 0
                for col_letter, fmt in spec.number_formats.items():
                    # Apply to all data rows (skip header row if present)
                    start = 2 if spec.headers else 1
                    for row_idx in range(start, max_row + 1):
                        ws[f"{col_letter}{row_idx}"].number_format = fmt

            # -- Individual CellSpec entries --
            for cell_spec in spec.cells:
                cell = ws.cell(row=cell_spec.row, column=cell_spec.col)
                if cell_spec.formula:
                    cell.value = cell_spec.formula
                elif cell_spec.value is not None:
                    cell.value = cell_spec.value
                if cell_spec.number_format:
                    cell.number_format = cell_spec.number_format
                if cell_spec.style:
                    font, fill, alignment = self._make_openpyxl_style(
                        cell_spec.style, Font, PatternFill, Alignment,
                    )
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if alignment:
                        cell.alignment = alignment

            # -- Chart --
            if spec.chart:
                self._add_chart(ws, spec.chart)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # -- Static helper methods -------------------------------------------------

    @staticmethod
    def _make_openpyxl_style(
        style: Any,
        font_cls: type,
        fill_cls: type,
        alignment_cls: type,
    ) -> tuple[Any, Any, Any]:
        """Convert a :class:`TextStyle` to openpyxl ``Font``, ``PatternFill``, ``Alignment``.

        Returns ``(font_or_None, fill_or_None, alignment_or_None)``.
        """
        font = None
        fill = None
        alignment = None

        font_kwargs: dict[str, Any] = {}
        if style.font_name:
            font_kwargs["name"] = style.font_name
        if style.font_size > 0:
            font_kwargs["size"] = style.font_size
        if style.bold:
            font_kwargs["bold"] = True
        if style.italic:
            font_kwargs["italic"] = True
        if style.color:
            hex_color = style.color.lstrip("#")
            if len(hex_color) == 6:
                font_kwargs["color"] = hex_color
        if font_kwargs:
            font = font_cls(**font_kwargs)

        # TextStyle doesn't carry fill info, but we prepare the slot
        # for future extension.

        if style.alignment and style.alignment != "left":
            alignment = alignment_cls(horizontal=style.alignment)

        return font, fill, alignment

    @staticmethod
    def _add_chart(ws: Any, chart_spec: Any) -> None:
        """Render a native openpyxl chart onto *ws*."""
        from firefly_dworkers.design.charts import ChartRenderer
        from firefly_dworkers.design.models import DataSeries, ResolvedChart

        # Normalise to ResolvedChart
        if isinstance(chart_spec, ResolvedChart):
            resolved = chart_spec
        elif isinstance(chart_spec, dict):
            resolved = ResolvedChart.model_validate(chart_spec)
        else:
            series_raw = getattr(chart_spec, "series", [])
            series = [
                DataSeries(**s) if isinstance(s, dict) else s for s in series_raw
            ]
            resolved = ResolvedChart(
                chart_type=getattr(chart_spec, "chart_type", "bar"),
                title=getattr(chart_spec, "title", ""),
                categories=getattr(chart_spec, "categories", []),
                series=series,
                colors=getattr(chart_spec, "colors", []),
                show_legend=getattr(chart_spec, "show_legend", True),
                show_data_labels=getattr(chart_spec, "show_data_labels", False),
                stacked=getattr(chart_spec, "stacked", False),
            )

        renderer = ChartRenderer()
        try:
            renderer.render_for_xlsx(resolved, ws)
        except (ImportError, ValueError) as exc:
            logger.warning("Failed to render chart to xlsx: %s", exc, exc_info=True)

    def _modify_sync(self, source: str, operations: list[SpreadsheetOperation]) -> bytes:
        wb = openpyxl.load_workbook(source)

        for op in operations:
            if op.operation == "add_sheet":
                name = op.data.get("name", "Sheet")
                wb.create_sheet(title=name)
            elif op.operation == "add_rows":
                ws_name = op.sheet_name or (wb.active.title if wb.active else "")
                ws = wb[ws_name]
                for row in op.data.get("rows", []):
                    ws.append(row)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
