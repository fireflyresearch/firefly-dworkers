"""SpreadsheetTool — parse CSV and Excel files.

CSV parsing uses the standard library.  Excel support requires ``openpyxl``
(install with ``pip install firefly-dworkers[data]``).
"""

from __future__ import annotations

import csv
import io
from collections.abc import Sequence
from typing import Any

from fireflyframework_genai.tools.base import BaseTool, GuardProtocol, ParameterSpec

from firefly_dworkers.tools.registry import tool_registry

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    OPENPYXL_AVAILABLE = False


@tool_registry.register("spreadsheet", category="data")
class SpreadsheetTool(BaseTool):
    """Parse and extract data from CSV and Excel files.

    Configuration parameters:

    * ``delimiter`` -- CSV delimiter character (default ``,``).
    * ``max_rows`` -- Default maximum rows to return.
    * ``encoding`` -- Text encoding for CSV data (default ``utf-8``).
    """

    def __init__(
        self,
        *,
        delimiter: str = ",",
        max_rows: int = 100,
        encoding: str = "utf-8",
        guards: Sequence[GuardProtocol] = (),
    ):
        super().__init__(
            "spreadsheet",
            description="Parse CSV and Excel files to extract structured data",
            tags=["data", "csv", "excel", "spreadsheet"],
            guards=guards,
            parameters=[
                ParameterSpec(
                    name="action",
                    type_annotation="str",
                    description="One of: parse_csv, parse_excel, describe, to_csv",
                    required=True,
                ),
                ParameterSpec(
                    name="content",
                    type_annotation="str",
                    description="Raw CSV content to parse (for CSV actions)",
                    required=False,
                    default="",
                ),
                ParameterSpec(
                    name="file_path",
                    type_annotation="str",
                    description="Path to an Excel file (for parse_excel action)",
                    required=False,
                    default="",
                ),
                ParameterSpec(
                    name="sheet_name",
                    type_annotation="str",
                    description="Excel sheet name (defaults to first sheet)",
                    required=False,
                    default="",
                ),
                ParameterSpec(
                    name="max_rows",
                    type_annotation="int",
                    description="Maximum rows to return",
                    required=False,
                    default=max_rows,
                ),
                ParameterSpec(
                    name="delimiter",
                    type_annotation="str",
                    description="CSV delimiter character",
                    required=False,
                    default=delimiter,
                ),
            ],
        )
        self._delimiter = delimiter
        self._max_rows = max_rows
        self._encoding = encoding

    async def _execute(self, **kwargs: Any) -> Any:
        action = kwargs["action"]
        content = kwargs.get("content", "")
        max_rows = kwargs.get("max_rows", self._max_rows)
        delimiter = kwargs.get("delimiter", self._delimiter)

        if action == "parse_csv":
            return self._parse_csv(content, max_rows, delimiter)
        if action == "describe":
            return self._describe_csv(content, delimiter)
        if action == "parse_excel":
            return self._parse_excel(
                kwargs.get("file_path", ""),
                kwargs.get("sheet_name", ""),
                max_rows,
            )
        if action == "to_csv":
            return self._describe_csv(content, delimiter)
        raise ValueError(f"Unknown action '{action}'; expected parse_csv, parse_excel, describe, or to_csv")

    def _parse_csv(self, content: str, max_rows: int, delimiter: str) -> dict[str, Any]:
        """Parse CSV content into rows."""
        reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
        rows = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(dict(row))
        return {
            "rows": rows,
            "row_count": len(rows),
            "columns": list(rows[0].keys()) if rows else [],
            "truncated": len(rows) >= max_rows,
        }

    def _describe_csv(self, content: str, delimiter: str) -> dict[str, Any]:
        """Describe the structure of CSV content."""
        reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        headers = next(reader, [])
        row_count = sum(1 for _ in reader)
        return {
            "columns": headers,
            "column_count": len(headers),
            "row_count": row_count,
        }

    def _parse_excel(self, file_path: str, sheet_name: str, max_rows: int) -> dict[str, Any]:
        """Parse an Excel file into rows."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel support. "
                "Install with: pip install firefly-dworkers[data]"
            )
        if not file_path:
            raise ValueError("parse_excel requires file_path")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        rows: list[dict[str, Any]] = []
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell) if cell is not None else f"col_{j}" for j, cell in enumerate(row)]
                continue
            if len(rows) >= max_rows:
                break
            rows.append({headers[j]: cell for j, cell in enumerate(row) if j < len(headers)})

        wb.close()
        return {
            "rows": rows,
            "row_count": len(rows),
            "columns": headers,
            "sheets": wb.sheetnames if hasattr(wb, "sheetnames") else [],
            "truncated": len(rows) >= max_rows,
        }
