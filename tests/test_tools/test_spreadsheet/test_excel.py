"""Tests for ExcelTool adapter."""

from __future__ import annotations

import io
import os
import tempfile

import pytest
from fireflyframework_genai.tools.base import BaseTool

from firefly_dworkers.design.models import DataSeries, ResolvedChart, TextStyle
from firefly_dworkers.tools.registry import tool_registry
from firefly_dworkers.tools.spreadsheet.base import SpreadsheetPort
from firefly_dworkers.tools.spreadsheet.excel import ExcelTool
from firefly_dworkers.tools.spreadsheet.models import CellSpec, SheetSpec, SpreadsheetOperation


class TestExcelToolRegistration:
    def test_is_spreadsheet_port(self) -> None:
        assert issubclass(ExcelTool, SpreadsheetPort)

    def test_is_base_tool(self) -> None:
        assert issubclass(ExcelTool, BaseTool)

    def test_registry_entry(self) -> None:
        assert tool_registry.has("excel")
        assert tool_registry.get_class("excel") is ExcelTool

    def test_category(self) -> None:
        assert tool_registry.get_category("excel") == "spreadsheet"

    def test_name(self) -> None:
        assert ExcelTool().name == "excel"


class TestExcelToolRead:
    async def test_read_spreadsheet(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")

        # Create a minimal .xlsx in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Value"])
        ws.append(["Alice", 100])
        ws.append(["Bob", 200])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.read())
            tmp_path = f.name

        try:
            tool = ExcelTool()
            result = await tool.execute(action="read", source=tmp_path)
            assert "sheets" in result
            assert len(result["sheets"]) == 1
            assert result["sheets"][0]["name"] == "Data"
            assert result["sheets"][0]["headers"] == ["Name", "Value"]
            assert len(result["sheets"][0]["rows"]) == 2
            assert result["sheets"][0]["rows"][0] == ["Alice", 100]
            assert result["sheets"][0]["rows"][1] == ["Bob", 200]
            assert result["active_sheet"] == "Data"
        finally:
            os.unlink(tmp_path)

    async def test_read_specific_sheet(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A", "B"])
        ws1.append([1, 2])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["X", "Y"])
        ws2.append([10, 20])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.read())
            tmp_path = f.name

        try:
            tool = ExcelTool()
            result = await tool.execute(action="read", source=tmp_path, sheet_name="Sheet2")
            assert len(result["sheets"]) == 1
            assert result["sheets"][0]["name"] == "Sheet2"
            assert result["sheets"][0]["headers"] == ["X", "Y"]
        finally:
            os.unlink(tmp_path)


class TestExcelToolCreate:
    async def test_create_spreadsheet_basic(self) -> None:
        pytest.importorskip("openpyxl")
        tool = ExcelTool()
        sheets = [
            SheetSpec(name="Revenue", headers=["Q1", "Q2"], rows=[[100, 200]]).model_dump(),
            SheetSpec(name="Expenses", headers=["Q1", "Q2"], rows=[[50, 75]]).model_dump(),
        ]
        result = await tool.execute(action="create", sheets=sheets)
        assert result["success"] is True
        assert result["bytes_length"] > 0

    async def test_create_spreadsheet_roundtrip(self) -> None:
        pytest.importorskip("openpyxl")
        tool = ExcelTool()

        # Create via the tool's internal method to get bytes
        sheets = [SheetSpec(name="Data", headers=["Name", "Value"], rows=[["Alice", 1]])]
        data = await tool._create_spreadsheet(sheets)

        # Write to temp and read back
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(data)
            tmp_path = f.name

        try:
            result = await tool.execute(action="read", source=tmp_path)
            assert result["sheets"][0]["name"] == "Data"
            assert result["sheets"][0]["headers"] == ["Name", "Value"]
            assert result["sheets"][0]["rows"][0] == ["Alice", 1]
        finally:
            os.unlink(tmp_path)

    async def test_create_multiple_sheets(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        sheets = [
            SheetSpec(name="Sheet1", headers=["A"], rows=[[1]]),
            SheetSpec(name="Sheet2", headers=["B"], rows=[[2]]),
        ]
        data = await tool._create_spreadsheet(sheets)

        # Verify by loading with openpyxl directly
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames
        wb.close()


class TestExcelToolModify:
    async def test_modify_add_rows(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")

        # Create source workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Value"])
        ws.append(["Alice", 100])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.read())
            tmp_path = f.name

        try:
            tool = ExcelTool()
            result = await tool.execute(
                action="modify",
                source=tmp_path,
                operations=[
                    {
                        "operation": "add_rows",
                        "sheet_name": "Data",
                        "data": {"rows": [["Bob", 200], ["Charlie", 300]]},
                    }
                ],
            )
            assert result["success"] is True
            assert result["bytes_length"] > 0
        finally:
            os.unlink(tmp_path)

    async def test_modify_add_sheet(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")

        # Create source workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.read())
            tmp_path = f.name

        try:
            tool = ExcelTool()
            data = await tool._modify_spreadsheet(
                tmp_path,
                [
                    SpreadsheetOperation(
                        operation="add_sheet",
                        data={"name": "NewSheet"},
                    )
                ],
            )

            # Verify the new sheet exists
            result_wb = openpyxl.load_workbook(io.BytesIO(data))
            assert "NewSheet" in result_wb.sheetnames
            result_wb.close()
        finally:
            os.unlink(tmp_path)


# ── New tests for Task 10: charts, formulas, styling, widths, number formats ──


class TestExcelToolBackwardsCompatibility:
    """Ensure the new fields on SheetSpec have defaults and don't break old code."""

    def test_sheet_spec_defaults(self) -> None:
        """SheetSpec with only name should still work (all new fields optional)."""
        spec = SheetSpec(name="Test")
        assert spec.chart is None
        assert spec.cells == []
        assert spec.header_style is None
        assert spec.cell_style is None
        assert spec.column_widths == []
        assert spec.number_formats == {}

    async def test_create_without_new_fields(self) -> None:
        """Creating with plain SheetSpec (no styling/charts) still works."""
        pytest.importorskip("openpyxl")
        tool = ExcelTool()
        sheets = [SheetSpec(name="Plain", headers=["A", "B"], rows=[[1, 2], [3, 4]])]
        data = await tool._create_spreadsheet(sheets)
        assert len(data) > 0

        import openpyxl as xl
        wb = xl.load_workbook(io.BytesIO(data))
        ws = wb["Plain"]
        assert ws.cell(1, 1).value == "A"
        assert ws.cell(2, 1).value == 1
        wb.close()


class TestExcelToolHeaderStyling:
    """Test header_style application to the header row."""

    async def test_header_bold_and_font(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        style = TextStyle(font_name="Arial", font_size=14, bold=True, color="#1A73E8")
        sheets = [
            SheetSpec(
                name="Styled",
                headers=["Name", "Revenue"],
                rows=[["Alice", 100]],
                header_style=style,
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Styled"]

        # Check header cell A1
        cell_a1 = ws.cell(1, 1)
        assert cell_a1.font.bold is True
        assert cell_a1.font.name == "Arial"
        assert cell_a1.font.size == 14
        # openpyxl stores color as Color object; check hex value
        assert cell_a1.font.color.rgb == "001A73E8"
        wb.close()

    async def test_header_italic(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        style = TextStyle(italic=True)
        sheets = [
            SheetSpec(
                name="Italic",
                headers=["Col1"],
                rows=[[42]],
                header_style=style,
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Italic"]
        assert ws.cell(1, 1).font.italic is True
        wb.close()

    async def test_header_alignment_center(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        style = TextStyle(alignment="center")
        sheets = [
            SheetSpec(
                name="Centered",
                headers=["H1", "H2"],
                rows=[[1, 2]],
                header_style=style,
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Centered"]
        assert ws.cell(1, 1).alignment.horizontal == "center"
        assert ws.cell(1, 2).alignment.horizontal == "center"
        wb.close()


class TestExcelToolCellStyling:
    """Test cell_style application to data rows."""

    async def test_cell_style_applied_to_data_rows(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        style = TextStyle(font_name="Courier", bold=True, color="#FF0000")
        sheets = [
            SheetSpec(
                name="DataStyle",
                headers=["X", "Y"],
                rows=[[10, 20], [30, 40]],
                cell_style=style,
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["DataStyle"]

        # Data starts at row 2 (row 1 is headers)
        for row_idx in (2, 3):
            for col_idx in (1, 2):
                cell = ws.cell(row_idx, col_idx)
                assert cell.font.bold is True
                assert cell.font.name == "Courier"

        # Header row should NOT have cell_style applied
        header_cell = ws.cell(1, 1)
        assert header_cell.font.name != "Courier" or not header_cell.font.bold
        wb.close()

    async def test_cell_style_without_headers(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        style = TextStyle(italic=True)
        sheets = [
            SheetSpec(
                name="NoHeaders",
                rows=[[1, 2], [3, 4]],
                cell_style=style,
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["NoHeaders"]
        # Data starts at row 1 when no headers
        assert ws.cell(1, 1).font.italic is True
        assert ws.cell(2, 2).font.italic is True
        wb.close()


class TestExcelToolColumnWidths:
    """Test column width application."""

    async def test_column_widths(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        sheets = [
            SheetSpec(
                name="Widths",
                headers=["Narrow", "Wide"],
                rows=[[1, 2]],
                column_widths=[10.0, 30.0],
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Widths"]
        assert ws.column_dimensions["A"].width == 10.0
        assert ws.column_dimensions["B"].width == 30.0
        wb.close()

    async def test_partial_column_widths(self) -> None:
        """Only first N columns get widths if list is shorter than header count."""
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        sheets = [
            SheetSpec(
                name="Partial",
                headers=["A", "B", "C"],
                rows=[[1, 2, 3]],
                column_widths=[15.0],
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Partial"]
        assert ws.column_dimensions["A"].width == 15.0
        # B and C use default
        wb.close()


class TestExcelToolNumberFormats:
    """Test number format application."""

    async def test_number_formats(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        sheets = [
            SheetSpec(
                name="Formats",
                headers=["Item", "Price", "Pct"],
                rows=[["Widget", 1234.5, 0.15], ["Gadget", 9876.54, 0.85]],
                number_formats={"B": "#,##0.00", "C": "0.0%"},
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Formats"]

        # Row 2 and 3 are data rows (row 1 is headers)
        assert ws["B2"].number_format == "#,##0.00"
        assert ws["B3"].number_format == "#,##0.00"
        assert ws["C2"].number_format == "0.0%"
        assert ws["C3"].number_format == "0.0%"
        wb.close()

    async def test_number_formats_without_headers(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        sheets = [
            SheetSpec(
                name="NoHead",
                rows=[[100, 0.5], [200, 0.75]],
                number_formats={"A": "#,##0", "B": "0%"},
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["NoHead"]
        assert ws["A1"].number_format == "#,##0"
        assert ws["B2"].number_format == "0%"
        wb.close()


class TestExcelToolFormulas:
    """Test formula writing via CellSpec."""

    async def test_formula_cell(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        sheets = [
            SheetSpec(
                name="Formulas",
                headers=["A", "B", "Sum"],
                rows=[[10, 20, 0], [30, 40, 0]],
                cells=[
                    CellSpec(row=2, col=3, formula="=A2+B2"),
                    CellSpec(row=3, col=3, formula="=A3+B3"),
                ],
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Formulas"]
        assert ws.cell(2, 3).value == "=A2+B2"
        assert ws.cell(3, 3).value == "=A3+B3"
        wb.close()

    async def test_cell_spec_value(self) -> None:
        """CellSpec can also write plain values."""
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        sheets = [
            SheetSpec(
                name="Values",
                headers=["Col"],
                rows=[[0]],
                cells=[
                    CellSpec(row=2, col=1, value=999),
                ],
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Values"]
        # CellSpec value overwrites the row data
        assert ws.cell(2, 1).value == 999
        wb.close()

    async def test_cell_spec_formula_takes_precedence_over_value(self) -> None:
        """When both formula and value are set, formula wins."""
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        sheets = [
            SheetSpec(
                name="Precedence",
                rows=[[0]],
                cells=[
                    CellSpec(row=1, col=1, value=42, formula="=1+1"),
                ],
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Precedence"]
        assert ws.cell(1, 1).value == "=1+1"
        wb.close()


class TestExcelToolCellSpecStyling:
    """Test individual CellSpec styling and number formats."""

    async def test_cell_spec_style(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        style = TextStyle(bold=True, color="#00FF00")
        sheets = [
            SheetSpec(
                name="CellStyle",
                headers=["A", "B"],
                rows=[[1, 2]],
                cells=[
                    CellSpec(row=2, col=1, style=style),
                ],
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["CellStyle"]
        cell = ws.cell(2, 1)
        assert cell.font.bold is True
        assert cell.font.color.rgb == "0000FF00"
        # cell B2 should not have the style
        cell_b2 = ws.cell(2, 2)
        assert cell_b2.font.bold is not True or cell_b2.font.color is None or cell_b2.font.color.rgb != "0000FF00"
        wb.close()

    async def test_cell_spec_number_format(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        sheets = [
            SheetSpec(
                name="CellFmt",
                rows=[[1234.5]],
                cells=[
                    CellSpec(row=1, col=1, number_format="$#,##0.00"),
                ],
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["CellFmt"]
        assert ws.cell(1, 1).number_format == "$#,##0.00"
        wb.close()


class TestExcelToolCharts:
    """Test chart embedding via ChartRenderer.render_for_xlsx()."""

    async def test_chart_from_resolved_chart(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        chart = ResolvedChart(
            chart_type="bar",
            title="Revenue",
            categories=["Q1", "Q2", "Q3"],
            series=[DataSeries(name="Sales", values=[100, 200, 300])],
        )
        sheets = [
            SheetSpec(
                name="WithChart",
                headers=["Quarter", "Sales"],
                rows=[["Q1", 100], ["Q2", 200], ["Q3", 300]],
                chart=chart,
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["WithChart"]
        # openpyxl stores charts in the worksheet's _charts list
        assert len(ws._charts) == 1
        # After round-trip, title is a Title object; extract text from rich text
        title_obj = ws._charts[0].title
        assert title_obj is not None
        # Extract the plain text from the Title's rich text paragraphs
        title_text = "".join(
            run.t for p in title_obj.tx.rich.p for run in (p.r or [])
        )
        assert title_text == "Revenue"
        wb.close()

    async def test_chart_from_dict(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        chart_dict = {
            "chart_type": "line",
            "title": "Trend",
            "categories": ["Jan", "Feb"],
            "series": [{"name": "Revenue", "values": [10, 20]}],
        }
        sheets = [
            SheetSpec(
                name="DictChart",
                headers=["Month", "Rev"],
                rows=[["Jan", 10], ["Feb", 20]],
                chart=chart_dict,
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["DictChart"]
        assert len(ws._charts) == 1
        title_obj = ws._charts[0].title
        assert title_obj is not None
        title_text = "".join(
            run.t for p in title_obj.tx.rich.p for run in (p.r or [])
        )
        assert title_text == "Trend"
        wb.close()

    async def test_no_chart_when_none(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        sheets = [
            SheetSpec(
                name="NoChart",
                headers=["A"],
                rows=[[1]],
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["NoChart"]
        assert len(ws._charts) == 0
        wb.close()


class TestExcelToolCombined:
    """Test combining multiple features in a single sheet."""

    async def test_all_features_together(self) -> None:
        openpyxl = pytest.importorskip("openpyxl")
        tool = ExcelTool()

        header_style = TextStyle(font_name="Arial", font_size=12, bold=True, color="#000000")
        cell_style = TextStyle(font_name="Calibri", font_size=10)
        chart = ResolvedChart(
            chart_type="bar",
            title="Combined",
            categories=["A", "B"],
            series=[DataSeries(name="Vals", values=[1, 2])],
        )

        sheets = [
            SheetSpec(
                name="Full",
                headers=["Item", "Price", "Qty"],
                rows=[["Widget", 9.99, 5], ["Gadget", 19.99, 3]],
                header_style=header_style,
                cell_style=cell_style,
                column_widths=[20.0, 15.0, 10.0],
                number_formats={"B": "#,##0.00"},
                chart=chart,
                cells=[
                    CellSpec(row=4, col=1, value="Total"),
                    CellSpec(row=4, col=2, formula="=SUM(B2:B3)"),
                ],
            ),
        ]
        data = await tool._create_spreadsheet(sheets)

        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Full"]

        # Headers styled
        assert ws.cell(1, 1).font.bold is True
        assert ws.cell(1, 1).font.name == "Arial"

        # Data rows styled
        assert ws.cell(2, 1).font.name == "Calibri"

        # Column widths
        assert ws.column_dimensions["A"].width == 20.0
        assert ws.column_dimensions["B"].width == 15.0
        assert ws.column_dimensions["C"].width == 10.0

        # Number formats
        assert ws["B2"].number_format == "#,##0.00"

        # CellSpec entries
        assert ws.cell(4, 1).value == "Total"
        assert ws.cell(4, 2).value == "=SUM(B2:B3)"

        # Chart
        assert len(ws._charts) == 1

        wb.close()


class TestMakeOpenpyxlStyle:
    """Unit tests for ExcelTool._make_openpyxl_style()."""

    def test_full_style(self) -> None:
        pytest.importorskip("openpyxl")
        from openpyxl.styles import Alignment, Font, PatternFill

        style = TextStyle(
            font_name="Helvetica",
            font_size=16,
            bold=True,
            italic=True,
            color="#AABBCC",
            alignment="center",
        )
        font, fill, alignment = ExcelTool._make_openpyxl_style(style, Font, PatternFill, Alignment)
        assert font is not None
        assert font.name == "Helvetica"
        assert font.size == 16
        assert font.bold is True
        assert font.italic is True
        assert font.color.rgb == "00AABBCC"
        assert fill is None  # TextStyle doesn't carry fill info
        assert alignment is not None
        assert alignment.horizontal == "center"

    def test_minimal_style(self) -> None:
        """TextStyle with all defaults produces no Font/Alignment."""
        pytest.importorskip("openpyxl")
        from openpyxl.styles import Alignment, Font, PatternFill

        style = TextStyle()
        font, fill, alignment = ExcelTool._make_openpyxl_style(style, Font, PatternFill, Alignment)
        assert font is None
        assert fill is None
        assert alignment is None

    def test_left_alignment_is_default(self) -> None:
        """alignment='left' should not create an Alignment object (it's the default)."""
        pytest.importorskip("openpyxl")
        from openpyxl.styles import Alignment, Font, PatternFill

        style = TextStyle(alignment="left")
        _, _, alignment = ExcelTool._make_openpyxl_style(style, Font, PatternFill, Alignment)
        assert alignment is None

    def test_color_without_hash(self) -> None:
        pytest.importorskip("openpyxl")
        from openpyxl.styles import Alignment, Font, PatternFill

        style = TextStyle(color="FF5500")
        font, _, _ = ExcelTool._make_openpyxl_style(style, Font, PatternFill, Alignment)
        assert font is not None
        assert font.color.rgb == "00FF5500"
